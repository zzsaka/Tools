#!/usr/bin/env python3
"""
MD to Excel Converter

このスクリプトはMarkdownファイルを解析してExcelファイルに変換します。
Markdownの見出し（#）は階層構造として解釈し、表（|）は対応するExcelの表として変換されます。
また、箇条書き（-、*、+）もサポートしています。

更新: 2025年5月11日
- エラーハンドリングの強化
- 日本語パス対応の改善
- テキスト処理の最適化
"""

import os
import re
import sys
import argparse
import logging
import traceback
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ロガー設定
logger = logging.getLogger(__name__)

# 例外をキャッチしてロギングするデコレータ
def log_exceptions(func):
    """
    関数の例外をキャッチしてログに記録するデコレータ
    """
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logger.error(f"{func.__name__}でエラーが発生しました: {str(e)}")
            logger.debug("詳細なエラー情報:", exc_info=True)
            raise
    return wrapper


class MdToExcelConverter:
    VERSION = "1.2.0"  # バージョン情報を更新
    
    def __init__(self, md_file, output_file=None, sheet_name="Sheet1", debug=False):
        """
        初期化メソッド
        
        Parameters:
            md_file (str): 入力Markdownファイルのパス
            output_file (str, optional): 出力Excelファイルのパス。デフォルトはNone（MDファイル名から自動生成）
            sheet_name (str, optional): 出力Excelファイルのシート名。デフォルトは"Sheet1"
            debug (bool, optional): デバッグモードを有効にするかどうか。デフォルトはFalse
        """
        # デバッグモードの設定
        if debug:
            logger.setLevel(logging.DEBUG)
            logger.debug("デバッグモードが有効です")
        
        self.md_file = md_file
        # 出力ファイル名が指定されていない場合は入力ファイル名から生成
        if output_file is None:
            md_path = Path(md_file)
            self.output_file = str(md_path.with_suffix('.xlsx'))
        else:
            self.output_file = output_file
            
        # 出力パスの検証
        output_path = Path(self.output_file)
        try:
            # 親ディレクトリが存在するか確認
            if not output_path.parent.exists():
                logger.info(f"出力ディレクトリ '{output_path.parent}' が存在しないため作成します")
                output_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.error(f"出力ディレクトリの作成中にエラーが発生しました: {str(e)}")
            raise
        
        self.sheet_name = sheet_name
        self.sections = []  # 各セクション（見出し、段落、表など）を格納
        self.current_row = 1
        self.debug = debug
        
        # バージョンと実行情報を表示
        logger.info(f"MD to Excel Converter v{self.VERSION}")
        logger.info(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"入力ファイル: {self.md_file}")
        logger.info(f"出力ファイル: {self.output_file}")
        logger.info(f"シート名: {self.sheet_name}")
        logger.info("-" * 50)

    @log_exceptions
    def parse_markdown(self):
        """
        Markdownファイルを解析し、セクション構造として抽出する
        """
        logger.info(f"Markdownファイルを解析中: {self.md_file}")
        
        # ファイルパスの検証
        md_path = Path(self.md_file)
        if not md_path.exists():
            raise FileNotFoundError(f"ファイル '{self.md_file}' が見つかりません。")
        if not md_path.is_file():
            raise ValueError(f"'{self.md_file}' はファイルではありません。")
            
        # ファイルの読み込みを試行（エンコーディングの自動検出）
        encodings = ['utf-8', 'shift-jis', 'euc-jp', 'iso-2022-jp']
        lines = None
        
        for encoding in encodings:
            try:
                with open(self.md_file, 'r', encoding=encoding) as file:
                    lines = file.readlines()
                logger.debug(f"ファイルを正常に読み込みました（エンコーディング: {encoding}）")
                break
            except UnicodeDecodeError:
                logger.debug(f"エンコーディング {encoding} では読み込めませんでした。別のエンコーディングを試します。")
                continue
            except Exception as e:
                logger.error(f"ファイル '{self.md_file}' を開けませんでした: {str(e)}")
                raise
        
        if lines is None:
            raise UnicodeError(f"ファイル '{self.md_file}' は対応しているエンコーディングで読み込めませんでした。UTF-8で保存し直してください。")

        current_section = None
        current_heading_level = 0
        in_table = False
        current_table = []
        paragraph_lines = []
        in_list = False
        current_list = []
        
        for line in lines:
            line = line.rstrip()
            
            # 見出しの検出
            heading_match = re.match(r'^(#{1,6})\s+(.+)$', line)
            if heading_match:
                # 前の段落を処理
                if paragraph_lines:
                    if current_section:
                        current_section['paragraphs'].append('\n'.join(paragraph_lines))
                    paragraph_lines = []
                
                # 前のリストを処理
                if in_list and current_list:
                    if current_section:
                        current_section['lists'].append(current_list)
                    current_list = []
                    in_list = False
                
                # 前の表を処理
                if in_table and current_table:
                    if current_section:
                        current_section['tables'].append(current_table)
                    current_table = []
                    in_table = False
                
                # 新しいセクションを作成
                level = len(heading_match.group(1))
                text = heading_match.group(2)
                
                current_section = {
                    'heading': text,
                    'level': level,
                    'paragraphs': [],
                    'lists': [],
                    'tables': []
                }
                
                self.sections.append(current_section)
                current_heading_level = level
                continue
            
            # 表の処理
            if line.startswith('|') and line.endswith('|'):
                # 前の段落を処理
                if paragraph_lines:
                    if current_section:
                        current_section['paragraphs'].append('\n'.join(paragraph_lines))
                    paragraph_lines = []
                
                # 前のリストを処理
                if in_list and current_list:
                    if current_section:
                        current_section['lists'].append(current_list)
                    current_list = []
                    in_list = False
                
                if not in_table:
                    # 新しい表の開始
                    in_table = True
                    current_table = []
                
                # 表の行を追加
                cells = [cell.strip() for cell in line.strip('|').split('|')]
                
                # セパレータ行（---）をスキップするが、整列情報は保存
                if all(re.match(r'^[-:\s]+$', cell) for cell in cells):
                    # セパレータ行の中で整列情報を取得
                    align_info = []
                    for cell in cells:
                        if cell.startswith(':') and cell.endswith(':'):
                            align_info.append('center')
                        elif cell.endswith(':'):
                            align_info.append('right')
                        else:
                            align_info.append('left')
                    
                    # この表の整列情報として保存
                    if not hasattr(self, 'table_alignments'):
                        self.table_alignments = {}
                    
                    # 現在のテーブルの位置（セクション内のインデックス）を計算
                    if current_section:
                        table_index = len(current_section['tables'])
                        self.table_alignments[(id(current_section), table_index)] = align_info
                    
                    continue
                
                current_table.append(cells)
                continue
            elif in_table:
                # 表の終了
                if current_section and current_table:
                    current_section['tables'].append(current_table)
                current_table = []
                in_table = False
            
            # 箇条書きの処理
            list_match = re.match(r'^(\s*)[-*+]\s+(.+)$', line)
            if list_match:
                # 前の段落を処理
                if paragraph_lines:
                    if current_section:
                        current_section['paragraphs'].append('\n'.join(paragraph_lines))
                    paragraph_lines = []
                
                if not in_list:
                    # 新しいリストの開始
                    in_list = True
                    current_list = []
                
                # リストアイテムを追加
                indent = len(list_match.group(1))
                content = list_match.group(2)
                current_list.append((indent, content))
                continue
            elif in_list and line.strip():
                # リストの終了（空行でない別の内容があれば終了）
                if not re.match(r'^\s*[-*+]', line):
                    if current_section and current_list:
                        current_section['lists'].append(current_list)
                    current_list = []
                    in_list = False
                    
                    # 次の処理に続く（この行は通常のテキストとして扱う）
                else:
                    # まだリスト内の行である
                    continue
            
            # 空行の処理
            if not line.strip():
                # 前の段落を処理
                if paragraph_lines:
                    if current_section:
                        current_section['paragraphs'].append('\n'.join(paragraph_lines))
                    paragraph_lines = []
                
                # 前のリストを処理
                if in_list and current_list:
                    if current_section:
                        current_section['lists'].append(current_list)
                    current_list = []
                    in_list = False
                    
                continue
            
            # 通常のテキスト
            if not in_list:
                paragraph_lines.append(line)
        
        # 残りの段落、リスト、表を処理
        if paragraph_lines:
            if current_section:
                current_section['paragraphs'].append('\n'.join(paragraph_lines))
        
        if in_list and current_list:
            if current_section:
                current_section['lists'].append(current_list)
        
        if in_table and current_table:
            if current_section:
                current_section['tables'].append(current_table)
        
        # 結果を出力
        print(f"セクション数: {len(self.sections)}")
        for i, section in enumerate(self.sections):
            print(f"  セクション {i+1}: {section['heading']} (レベル {section['level']})")
            print(f"    段落数: {len(section['paragraphs'])}")
            print(f"    リスト数: {len(section['lists'])}")
            print(f"    表数: {len(section['tables'])}")

    def _get_column_width(self, text):
        """
        テキストの長さに基づいてカラム幅を計算する
        全角文字（日本語など）と半角文字の幅を適切に計算
        
        Parameters:
            text (str): セルのテキスト
            
        Returns:
            float: 推奨されるカラム幅
        """
        if text is None:
            return 10.0
            
        # 文字列に変換
        text = str(text)
        
        # 文字の種類ごとに幅を計算
        # 1. 日本語文字（漢字、ひらがな、カタカナ）: 2.0幅
        japanese_chars = len(re.findall(r'[\u3000-\u303f\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fff]', text))
        
        # 2. 全角英数字・記号: 2.0幅
        fullwidth_chars = len(re.findall(r'[\uff01-\uff60]', text))
        
        # 3. 半角英数字・記号: 1.0幅
        halfwidth_chars = len(text) - japanese_chars - fullwidth_chars
        
        # 4. 改行文字も考慮（改行ごとに幅が変わる可能性）
        newlines = text.count('\n')
        
        # 基本幅 + 各文字タイプの追加幅
        # 余裕を持たせるために係数1.2をかける
        width = 2.0 + (japanese_chars * 2.0 + fullwidth_chars * 2.0 + halfwidth_chars * 1.0) * 1.2
        
        # 改行がある場合は最大の行の長さを考慮
        if newlines > 0:
            lines = text.split('\n')
            max_line_width = max([self._get_column_width(line) for line in lines])
            width = max(width, max_line_width)
            
        return min(max(width, 10.0), 60.0)  # 最小10、最大60の幅に制限

    @log_exceptions
    def create_excel(self):
        """
        解析したMarkdownの内容からExcelファイルを生成する
        """
        logger.info(f"Excelファイルを作成中: {self.output_file}")
        
        # Excelワークブックの作成
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_name
        
        # ワークブック全体のデフォルトフォントをメイリオに設定
        # 注：openpyxlの制限により、これはワークブックレベルではなく各セルに適用する必要がある
        
        # スタイルの定義
        heading_fonts = {
            1: Font(name='メイリオ', bold=True, size=16, color="000000"),
            2: Font(name='メイリオ', bold=True, size=14, color="000000"),
            3: Font(name='メイリオ', bold=True, size=12, color="000000"),
            4: Font(name='メイリオ', bold=True, size=11, color="000000"),
            5: Font(name='メイリオ', bold=True, size=10, color="000000"),
            6: Font(name='メイリオ', bold=True, size=10, color="000000")
        }
        
        heading_fills = {
            1: PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            2: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            3: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            4: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            5: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            6: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        }
        
        # 境界線スタイル
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダースタイル（表用）
        header_font = Font(name='メイリオ', bold=True)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # 中央揃え
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 箇条書きのスタイル
        list_font = Font(name='メイリオ', size=11)
        
        # 各セクションを追加
        for section in self.sections:
            # 見出しを追加
            sheet.cell(row=self.current_row, column=1, value=section['heading'])
            
            # 見出しのスタイル設定
            cell = sheet.cell(row=self.current_row, column=1)
            cell.font = heading_fonts.get(section['level'], Font(bold=True))
            cell.fill = heading_fills.get(section['level'], PatternFill())
            
            # インデント設定（レベルに応じて）
            cell.alignment = Alignment(indent=section['level']-1)
            
            # 見出しの行の高さを調整
            sheet.row_dimensions[self.current_row].height = 24
            
            self.current_row += 1
            
            # 段落を追加
            for paragraph in section['paragraphs']:
                if paragraph.strip():  # 空の段落はスキップ
                    cell = sheet.cell(row=self.current_row, column=1, value=paragraph)
                    cell.font = Font(name='メイリオ', size=11)
                    
                    # 改行を含む段落の場合は、セルの書式設定を調整
                    if '\n' in paragraph:
                        # 折り返し設定を有効にし、上下中央揃えに
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                        
                        # 行の高さを内容に合わせて自動調整（目安として改行1つにつき15ポイント加算）
                        line_count = paragraph.count('\n') + 1
                        row_height = max(24, 15 * line_count)  # 最低24、それ以上は改行数に応じて
                        sheet.row_dimensions[self.current_row].height = row_height
                    
                    self.current_row += 1
            
            # リストを追加
            for list_items in section['lists']:
                for indent, content in list_items:
                    # インデントレベルを計算（スペース4つを1レベルと仮定）
                    indent_level = max(1, indent // 2 + 1)
                    
                    # リストアイテムを追加
                    cell = sheet.cell(row=self.current_row, column=1, value=f"• {content}")
                    cell.font = list_font
                    cell.alignment = Alignment(indent=indent_level)
                    
                    self.current_row += 1
                
                # リストの後に空行を追加
                self.current_row += 1
            
            # 表を追加
            for table in section['tables']:
                if not table:  # 空の表はスキップ
                    continue
                    
                # ヘッダー行が存在するか確認
                if len(table) > 0:
                    # 表のヘッダー行
                    for col_index, header in enumerate(table[0], start=1):
                        cell = sheet.cell(row=self.current_row, column=col_index, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        cell.border = thin_border
                        
                        # カラム幅の調整
                        col_letter = get_column_letter(col_index)
                        sheet.column_dimensions[col_letter].width = self._get_column_width(header)
                    
                    self.current_row += 1
                    
                    # 表のデータ行
                    # 表の整列情報を取得（存在する場合）
                    table_align_info = None
                    if hasattr(self, 'table_alignments'):
                        table_align_info = self.table_alignments.get((id(section), section['tables'].index(table)), None)
                    
                    for row_data in table[1:]:
                        for col_index, cell_value in enumerate(row_data, start=1):
                            # 列インデックスが範囲外の場合は空白セルを追加
                            if col_index <= len(row_data):
                                cell = sheet.cell(row=self.current_row, column=col_index, value=cell_value)
                                cell.font = Font(name='メイリオ', size=11)
                                cell.border = thin_border
                                
                                # セルの配置を決定
                                horizontal_align = 'left'  # デフォルト
                                
                                # 1. Markdownの表の整列情報がある場合、そちらを優先
                                if table_align_info and col_index - 1 < len(table_align_info):
                                    horizontal_align = table_align_info[col_index - 1]
                                # 2. それ以外は、数値は右揃え、他は左揃え
                                else:
                                    try:
                                        # 金額など「,」を含む数値文字列のカンマを一時的に除去
                                        numeric_value = cell_value.replace(',', '') if isinstance(cell_value, str) else cell_value
                                        # 数値変換を試みる
                                        float(numeric_value)
                                        # 数値の場合は右揃え
                                        horizontal_align = 'right'
                                    except (ValueError, TypeError, AttributeError):
                                        # 数値でない場合は左揃え
                                        horizontal_align = 'left'
                                
                                # 配置を設定
                                cell.alignment = Alignment(horizontal=horizontal_align, vertical='center')
                                
                                # カラム幅の再調整
                                col_letter = get_column_letter(col_index)
                                current_width = sheet.column_dimensions[col_letter].width
                                new_width = self._get_column_width(cell_value)
                                sheet.column_dimensions[col_letter].width = max(current_width, new_width)
                        
                        self.current_row += 1
                
                # 表の後に空行を追加
                self.current_row += 1
        
        # ファイルの保存
        try:
            # 出力ファイルのディレクトリが存在するか確認
            output_dir = os.path.dirname(self.output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                logger.debug(f"出力ディレクトリを作成しました: {output_dir}")
            
            workbook.save(self.output_file)
            logger.info(f"ファイルを保存しました: {self.output_file}")
        except PermissionError:
            logger.error(f"エラー: ファイル '{self.output_file}' が他のプログラムで開かれているため保存できません。")
            logger.error("Excelを閉じてから再試行してください。")
            raise
        except Exception as e:
            logger.error(f"エラー: ファイルの保存中にエラーが発生しました: {str(e)}")
            if self.debug:
                logger.debug("詳細なエラー情報:", exc_info=True)
            raise

    @log_exceptions
    def convert(self):
        """
        変換プロセスを実行する
        """
        start_time = datetime.now()
        logger.debug(f"変換開始: {start_time.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")
        
        self.parse_markdown()
        self.create_excel()
        
        end_time = datetime.now()
        elapsed = end_time - start_time
        logger.debug(f"変換終了: {end_time.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")
        logger.debug(f"処理時間: {elapsed.total_seconds():.3f}秒")
        
        return self.output_file


def main():
    """
    メイン実行関数
    コマンドライン引数を解析して処理を実行する
    """
    parser = argparse.ArgumentParser(description="Markdownファイルをエクセルに変換するプログラム")
    parser.add_argument("input", help="入力Markdownファイルのパス", nargs='+')
    parser.add_argument("-o", "--output", help="出力Excelファイルのパス（指定しない場合は入力ファイル名と同じ名前で.xlsxとして保存）")
    parser.add_argument("-s", "--sheet", default="Sheet1", help="出力Excelファイルのシート名（デフォルト: Sheet1）")
    parser.add_argument("-d", "--debug", action="store_true", help="デバッグ情報を表示")
    parser.add_argument("-v", "--version", action="store_true", help="バージョン情報を表示して終了")
    parser.add_argument("--overwrite", action="store_true", help="既存のファイルを上書きする（警告を表示しない）")
    
    args = parser.parse_args()
    
    # ロギングを設定
    log_level = logging.DEBUG if args.debug else logging.INFO
    logging.basicConfig(
        level=log_level,
        format='%(levelname)s: %(message)s'
    )
    
    # バージョン情報の表示
    if args.version:
        print(f"MD to Excel Converter v{MdToExcelConverter.VERSION}")
        return
    
    # 出力ファイルが指定されているが、入力ファイルが複数ある場合は警告
    if args.output and len(args.input) > 1:
        logging.warning("複数の入力ファイルが指定されていますが、出力ファイルは1つだけ指定されています。")
        logging.warning("最初のファイルのみが指定された出力ファイルに変換され、他のファイルはデフォルトの名前で保存されます。")
    
    # 各入力ファイルを処理
    success_count = 0
    error_count = 0
    
    for i, input_file in enumerate(args.input):
        try:
            # 最初のファイルだけ指定された出力ファイル名を使用
            output_file = args.output if i == 0 and args.output else None
            
            # 既存のファイルを確認
            if output_file and os.path.exists(output_file) and not args.overwrite:
                response = input(f"警告: ファイル '{output_file}' は既に存在します。上書きしますか？ (y/n): ")
                if response.lower() != 'y':
                    logging.info(f"処理をスキップします: {input_file}")
                    continue
            
            # 変換実行
            converter = MdToExcelConverter(input_file, output_file, args.sheet)
            result_file = converter.convert()
            logging.info(f"変換完了: {result_file}")
            success_count += 1
            
        except Exception as e:
            logging.error(f"ファイル '{input_file}' の処理中にエラーが発生しました: {str(e)}")
            if args.debug:
                import traceback
                traceback.print_exc()
            error_count += 1
    
    # 処理結果の表示
    if len(args.input) > 1:
        logging.info(f"処理結果: {success_count}件成功, {error_count}件失敗 (合計: {len(args.input)}件)")


if __name__ == "__main__":
    main()